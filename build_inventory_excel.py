"""
이카운트 재고 → Excel 누적 저장 스크립트
─────────────────────────────────────────
매주 금요일 실행해 이카운트 전 창고 재고를 조회하고,
기존 Excel(재고현황.xlsx)에 새 날짜 데이터를 추가한다.

동작 요약:
  1) 이카운트에서 설정된 모든 창고 재고를 한 번에 조회
  2) 창고 그룹(온라인/영업/사업지원TF/컨기부)별로 품목 수량 집계
  3) 기존 Excel이 있으면:
     - 품목코드 기준으로 수량 업데이트
     - 통계 시트에 오늘 날짜 열 추가(또는 덮어쓰기)
     - 기존 입고단가 보존
  4) 기존 Excel이 없으면: 조회 결과로 새 파일 생성
  5) output/ 폴더에 날짜_재고현황.xlsx 저장

사용 예:
  python build_inventory_excel.py                              # 오늘 날짜 기준
  python build_inventory_excel.py --date 20260616             # 날짜 지정
  python build_inventory_excel.py --base 재고현황.xlsx         # 기존 파일 기준으로 업데이트
"""

import sys
import re
import json
import argparse
import traceback
import unicodedata
from pathlib import Path
from datetime import datetime, date
from collections import defaultdict

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ecount_inventory import get_inventory_by_location, extract_rows
from ecount_warehouses import WH_GROUPS, BRAND_SHEET_MAP, ALL_WH_CODES
from ecount_items import fetch_item_master

# 그룹코드 → 브랜드시트 열 헤더 표시명
def _to_int_price(v) -> int:
    """price 값을 int로 안전 변환. 빈 문자열·None·콤마 포함 문자열 모두 처리."""
    if v is None or v == "":
        return 0
    try:
        return int(float(str(v).replace(",", "")))
    except (ValueError, TypeError):
        return 0


GROUP_DISPLAY = {
    "온라인":    "온라인",
    "영업":      "영업",
    "영업_위탁": "영업_위탁",
    "사업지원TF": "사업지원",
    "컨기부":    "컨기",
}

# 그룹코드 → 통계 시트 섹션명
GROUP_SECTION_NAMES = {
    "온라인":    "온라인사업부",
    "영업":      "영업사업부",
    "영업_위탁": "영업사업부_위탁매장",
    "사업지원TF": "사업지원 TF",
    "컨기부":    "컨텐츠기획부",
}

# 통계 브랜드 표시 순서
STATS_BRAND_ORDER = [
    "굿스마일", "메가하우스", "부시로드", "블로키", "반다이남코",
    "핫토이", "하비재팬", "CCS TOYS", "P001", "코코파스튜디오", "마그넷", "기타브랜드",
]

# ─────────────────────────────────────────
# 브랜드 추출
# ─────────────────────────────────────────

# 이카운트 API에서 오는 다양한 브랜드명 표기 → 정규 브랜드명 매핑 (소문자 키)
_BRAND_NORM: dict[str, str] = {
    # 굿스마일
    "굿스마일컴퍼니": "굿스마일",
    "굿스마일(good smile)": "굿스마일",
    "굿스마일상하이": "굿스마일",
    "굿스마일아츠상하이": "굿스마일",
    "good smile company, inc. (굿스마일)": "굿스마일",
    "good smile": "굿스마일",
    "good smile company": "굿스마일",
    "goodsmilecompany": "굿스마일",
    "gsc": "굿스마일",
    "max factory": "굿스마일",
    "orange rouge": "굿스마일",
    "freeing": "굿스마일",
    "phat!": "굿스마일",
    "phat! company": "굿스마일",
    # 메가하우스
    "메가하우스(megahouse)": "메가하우스",
    "mega house (메가하우스)": "메가하우스",
    "megahouse": "메가하우스",
    # 부시로드
    "bushiroad creative": "부시로드",
    # 반다이남코
    "반다이": "반다이남코",
    "bandai namco arts": "반다이남코",
    "bandai namco filmworks": "반다이남코",
    "bandai": "반다이남코",
    # 핫토이
    "핫토이(hottoys)": "핫토이",
    "핫토이(hot toys)": "핫토이",
    "hot toys": "핫토이",
    "hottoys": "핫토이",
    # 하비재팬
    "hobby japan (하비재팬)": "하비재팬",
    "hobby japan": "하비재팬",
    "hobbyjapan": "하비재팬",
    # CCS TOYS
    "ccs": "CCS TOYS",
    "ccstoys": "CCS TOYS",
    # 코코파스튜디오
    "코코파": "코코파스튜디오",
}


def _nfc(s: str) -> str:
    return unicodedata.normalize("NFC", s.strip())


def normalize_brand_name(brand: str) -> str:
    """이카운트 브랜드명을 정규 브랜드명으로 변환한다."""
    if not brand:
        return "기타브랜드"
    cleaned = _nfc(brand)
    # 1) 정확히 NAMED_BRANDS에 있으면 바로 반환
    if cleaned in NAMED_BRANDS:
        return cleaned
    # 2) 소문자로 _BRAND_NORM 매핑 시도
    mapped = _BRAND_NORM.get(cleaned.lower())
    if mapped:
        return mapped
    # 3) NAMED_BRANDS에서 대소문자·공백 무시하고 탐색
    cl = cleaned.lower()
    for nb in NAMED_BRANDS:
        if _nfc(nb).lower() == cl:
            return nb
    return cleaned


BRAND_ALIASES = {
    "굿스마일컴퍼니": "굿스마일",
    "GSC": "굿스마일",
    "GOOD SMILE": "굿스마일",
    "MEGAHOUSE": "메가하우스",
    "BUSHIROAD": "부시로드",
    "BANDAI": "반다이남코",
    "HOT TOYS": "핫토이",
    "HOTTOYS": "핫토이",
    "HOBBY JAPAN": "하비재팬",
    "HOBBYJAPAN": "하비재팬",
}

KNOWN_BRANDS = list(BRAND_SHEET_MAP.keys())


def extract_brand(prod_des: str) -> str:
    """PROD_DES에서 브랜드명을 추출한다. '[브랜드명]...' 형식 우선."""
    if not prod_des:
        return "기타브랜드"
    m = re.match(r"\[([^\]]+)\]", prod_des.strip())
    if m:
        raw = _nfc(m.group(1).strip())  # NFC 정규화로 한글 NFD/NFC 불일치 방지
        for brand in KNOWN_BRANDS:
            if raw == _nfc(brand) or raw.lower() == _nfc(brand).lower():
                return brand
        for alias, canonical in BRAND_ALIASES.items():
            if _nfc(alias).lower() in raw.lower():
                return canonical
        return raw
    return "기타브랜드"


# ─────────────────────────────────────────
# ECOUNT 조회 + 집계
# ─────────────────────────────────────────

def fetch_all_groups(base_date: str) -> dict:
    """
    전체 창고 재고를 1회 조회 후 창고코드로 그룹별 집계한다.
    (창고별 개별 조회 → 10분 한도 소진 방지)

    반환: {
        prod_cd: {
            "prod_des": str,
            "brand": str,
            "groups": { "온라인": qty, "영업": qty, ... }
        }
    }
    """
    # 사용할 창고코드 집합 (그룹에 포함된 것만)
    wh_to_group: dict[str, str] = {}
    for group, codes in WH_GROUPS.items():
        for code in codes:
            wh_to_group[code] = group

    # 품목 마스터 로드 (입고단가 + 브랜드명 + 바코드)
    print("  품목 마스터 로드 중...")
    item_master, barcode_index = fetch_item_master()

    print("  전체 창고 재고 조회 중 (1회)...", end=" ", flush=True)
    try:
        raw = get_inventory_by_location(base_date)
    except Exception as e:
        if "412" in str(e) or "Precondition" in str(e):
            print(f"\n  세션 만료 감지, 강제 재로그인 후 재시도...")
            raw = get_inventory_by_location(base_date, force_login=True)
        else:
            raise
    all_rows = extract_rows(raw)
    print(f"{len(all_rows)}건")

    group_data: dict[str, dict[str, float]] = {g: defaultdict(float) for g in WH_GROUPS}
    # 재고 API 응답에서 품목명/브랜드 보조 정보 수집 (품목마스터에 없는 품목 대비)
    inv_prod_info: dict[str, dict] = {}

    for row in all_rows:
        if not isinstance(row, dict):
            continue

        wh_cd   = str(row.get("WH_CD") or "").strip()
        prod_cd = str(row.get("PROD_CD") or "").strip()
        if not prod_cd or wh_cd not in wh_to_group:
            continue

        # 재고 행에서 품목명·브랜드 수집 (한 번만 저장)
        if prod_cd not in inv_prod_info:
            inv_prod_info[prod_cd] = {
                "prod_des": str(row.get("PROD_DES") or "").strip(),
                "brand":    str(row.get("CONT1") or "").strip(),
            }

        group = wh_to_group[wh_cd]
        try:
            qty = float(str(row.get("BAL_QTY") or 0).replace(",", ""))
        except (ValueError, TypeError):
            qty = 0.0
        group_data[group][prod_cd] += qty

    # 그룹별 집계 현황 출력
    for g, gd in group_data.items():
        total_qty = sum(gd.values())
        print(f"  [{g}] {len(gd)}개 품목, 총 {int(total_qty):,}개")

    # 통합 (품목 마스터에서 단가·브랜드 보강)
    all_prod_cds: set[str] = set()
    for gd in group_data.values():
        all_prod_cds.update(gd.keys())

    result = {}
    for prod_cd in all_prod_cds:
        # 품목코드로 먼저 조회, 없으면 바코드 역방향 조회 (컨기부 등 바코드=품목코드 케이스)
        master = item_master.get(prod_cd) or barcode_index.get(prod_cd, {})
        # 품목마스터에 없으면 재고 API 응답에서 보완
        inv_info = inv_prod_info.get(prod_cd, {})
        prod_des = master.get("prod_des") or inv_info.get("prod_des") or prod_cd
        # 브랜드: 품목마스터 CONT1 → 재고API CONT1 → 품목명 파싱 순으로 시도
        _cont1 = master.get("brand") or inv_info.get("brand") or ""
        if _cont1 and _cont1 != "기타브랜드":
            brand = _cont1
        else:
            brand = extract_brand(prod_des)
        result[prod_cd] = {
            "prod_des": prod_des,
            "brand":    brand,
            "in_price": master.get("in_price", 0.0),
            "groups":   {g: group_data[g].get(prod_cd, 0.0) for g in WH_GROUPS},
        }

    return result


# ─────────────────────────────────────────
# Excel 로드 / 파싱
# ─────────────────────────────────────────

# 자체 시트를 갖는 브랜드 목록 (여기 없으면 기타브랜드 시트로 합산)
NAMED_BRANDS = {
    "굿스마일", "메가하우스", "부시로드", "블로키", "반다이남코",
    "핫토이", "하비재팬", "CCS TOYS", "CCS", "P001",
    "코코파", "코코파스튜디오", "마그넷",
}


def load_base_excel(path: Path) -> tuple[dict, dict]:
    """
    기존 Excel에서 가격 정보를 로드한다.

    반환:
      prod_map   : prod_cd  → {brand, name, price, sheet}  (품목코드 컬럼 있는 구형 포맷)
      name_price : 품목명   → price                         (브랜드|품목명|입고단가 신형 포맷)
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    prod_map   = {}
    name_price = {}

    for sh_name in wb.sheetnames:
        if sh_name == "통계":
            continue
        ws = wb[sh_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(h or "").strip() for h in rows[0]]

        # 신형 포맷: 브랜드 | 품목명 | 입고단가 | ...
        if len(header) >= 3 and header[0] == "브랜드" and header[1] == "품목명" and header[2] == "입고단가":
            for row in rows[1:]:
                if not row or row[1] is None:
                    continue
                name = str(row[1]).strip()
                price = row[2]
                if name and price not in (None, 0, ""):
                    name_price[name] = price

        # 구형 포맷: 브랜드 | 품목코드 | 품목명 | 입고단가 | ...
        elif len(header) >= 4 and header[1] == "품목코드":
            for row in rows[1:]:
                if not row or row[1] is None:
                    continue
                prod_cd = str(row[1]).strip()
                if not prod_cd:
                    continue
                prod_map[prod_cd] = {
                    "brand": str(row[0] or "").strip(),
                    "name":  str(row[2] or "").strip(),
                    "price": row[3],
                    "sheet": sh_name,
                }

        # 직영 포맷: 재고위치 | 바코드 | ...
        elif len(header) >= 4 and header[1] == "바코드":
            for row in rows[1:]:
                if not row or row[1] is None:
                    continue
                prod_cd = str(row[1]).strip()
                if not prod_cd:
                    continue
                prod_map[prod_cd] = {
                    "brand": sh_name,
                    "name":  str(row[2] or "").strip(),
                    "price": row[4] if len(row) > 4 else None,
                    "sheet": sh_name,
                }

    wb.close()
    return prod_map, name_price


# ─────────────────────────────────────────
# Excel 생성
# ─────────────────────────────────────────

# 스타일 상수
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
BRAND_FILL  = PatternFill("solid", fgColor="D9E1F2")
THIN = Side(style="thin", color="AAAAAA")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")


def _active_groups(rows: list, group_cols: list) -> list:
    """해당 브랜드 품목 중 수량이 있는 그룹만 반환한다."""
    active = []
    for g in group_cols:
        if any(item["groups"].get(g, 0) != 0 for item in rows):
            active.append(g)
    # 수량이 있는 그룹이 하나도 없으면 전체 반환
    return active if active else group_cols


def _write_brand_sheet(ws, brand: str, rows: list, group_cols: list):
    """
    브랜드별 시트를 작성한다.
    rows: [ {name, price, groups:{온라인:qty, ...}} ]
    group_cols: WH_GROUPS 키 목록 (데이터 있는 그룹만 동적으로 축소됨)
    """
    ws.freeze_panes = "A2"

    # 데이터 있는 그룹만 표시
    active = _active_groups(rows, group_cols)

    # 헤더: 브랜드, 품목명, 입고단가, (그룹별 수량/금액)
    # 표시명: 사업지원TF → 사업지원, 컨기부 → 컨기
    headers = ["브랜드", "품목명", "입고단가"]
    for g in active:
        disp = GROUP_DISPLAY.get(g, g)
        headers += [f"{disp}수량" if disp == "영업" else f"{disp} 수량",
                    f"{disp} 재고금액"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # 열 너비
    col_widths = [10, 55, 12] + [12, 16] * len(active)
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    price_col = 3

    # 데이터 행
    for r_idx, item in enumerate(rows, 2):
        ws.cell(row=r_idx, column=1, value=brand)
        ws.cell(row=r_idx, column=2, value=item["name"])
        ws.cell(row=r_idx, column=3, value=item["price"])

        col_offset = 4
        for g in active:
            qty = item["groups"].get(g, 0.0)
            qty_val = int(qty) if qty == int(qty) else qty
            price = _to_int_price(item["price"])
            amt_val = int(qty_val * price)
            ws.cell(row=r_idx, column=col_offset, value=qty_val)
            ws.cell(row=r_idx, column=col_offset + 1, value=amt_val)
            col_offset += 2

    # 합계 행
    total_row = len(rows) + 2
    ws.cell(row=total_row, column=2, value="합계").font = Font(bold=True)
    col_offset = 4
    for g in active:
        total_qty = sum(
            int(it["groups"].get(g, 0)) for it in rows
        )
        total_amt = sum(
            int(it["groups"].get(g, 0)) * _to_int_price(it["price"]) for it in rows
        )
        ws.cell(row=total_row, column=col_offset, value=total_qty).font = Font(bold=True)
        ws.cell(row=total_row, column=col_offset + 1, value=total_amt).font = Font(bold=True)
        col_offset += 2


def _find_date_col(ws_stats, header_row: int, ref_date: date) -> int:
    """
    섹션 header_row에서 ref_date 열 위치를 찾거나 새 열 위치를 반환한다.
    전체 max_column 대신 해당 행의 날짜 셀만 순서대로 탐색해
    섹션마다 열이 밀리는 문제를 방지한다.
    """
    last_date_col = 0
    col = 2
    while True:
        cv = ws_stats.cell(row=header_row, column=col).value
        if isinstance(cv, (date, datetime)):
            cd = cv.date() if isinstance(cv, datetime) else cv
            if cd == ref_date:
                return col          # 이미 존재하는 날짜 → 덮어쓰기
            last_date_col = col
            col += 2
        else:
            break                   # 날짜 없는 셀 → 이 섹션 날짜 끝
    # 마지막 날짜 열 바로 다음 짝수 열
    return (last_date_col + 2) if last_date_col else 2


def _write_section(ws, start_row: int, sect_name: str, ref_date: date,
                   group_data: dict, brand_order: list):
    """
    통계 시트에 섹션 한 개를 작성한다 (새 파일용).
    group_data: {brand: {수량, 금액}}
    반환: 다음 섹션 시작 행 번호
    """
    r = start_row

    # 섹션 헤더 행 (부서명 + 날짜)
    c = ws.cell(row=r, column=1, value=sect_name)
    c.font = Font(bold=True)
    ws.cell(row=r, column=2, value=ref_date).number_format = "M월 D일"
    r += 1

    # 수량/금액 서브헤더
    ws.cell(row=r, column=2, value="수량")
    ws.cell(row=r, column=3, value="금액")
    r += 1

    # 브랜드 행
    for brand in brand_order:
        d = group_data.get(brand)
        if d is None:
            continue
        ws.cell(row=r, column=1, value=brand)
        ws.cell(row=r, column=2, value=d["수량"])
        ws.cell(row=r, column=3, value=d["금액"])
        r += 1

    # 합 계 행
    total_qty = sum(d["수량"] for d in group_data.values())
    total_amt = sum(d["금액"] for d in group_data.values())
    ws.cell(row=r, column=1, value="합 계").font = Font(bold=True)
    ws.cell(row=r, column=2, value=total_qty).font = Font(bold=True)
    ws.cell(row=r, column=3, value=total_amt).font = Font(bold=True)
    r += 1

    # 빈 행 구분
    r += 1
    return r


def _update_section(ws_stats, sect_start: int, ref_date: date, group_data: dict):
    """
    기존 통계 섹션에 새 날짜 열을 추가한다.
    sect_start: 섹션 헤더 행 번호 (부서명이 있는 행)
    group_data에 있는 브랜드가 기존 행에 없으면 합 계 앞에 행을 삽입한다.
    """
    date_col = _find_date_col(ws_stats, sect_start, ref_date)

    ws_stats.cell(row=sect_start, column=date_col,
                  value=ref_date).number_format = "M월 D일"
    ws_stats.cell(row=sect_start + 1, column=date_col, value="수량")
    ws_stats.cell(row=sect_start + 1, column=date_col + 1, value="금액")

    # 1pass: 기존 브랜드 행에 값 채우고, 합 계 행 위치 기록
    written_brands = set()
    total_row = None
    for row in ws_stats.iter_rows(min_row=sect_start + 2):
        brand_cell = row[0]
        brand = str(brand_cell.value or "").strip()
        if not brand:
            break
        if brand == "합 계":
            total_row = brand_cell.row
            break
        d = group_data.get(brand, {"수량": 0, "금액": 0})
        ws_stats.cell(row=brand_cell.row, column=date_col, value=d["수량"])
        ws_stats.cell(row=brand_cell.row, column=date_col + 1, value=d["금액"])
        written_brands.add(brand)

    # 2pass: group_data에 있지만 기존 행에 없는 브랜드 → 합 계 앞에 삽입
    missing = [b for b in STATS_BRAND_ORDER if b in group_data and b not in written_brands]
    if missing and total_row:
        ws_stats.insert_rows(total_row, len(missing))
        for i, brand in enumerate(missing):
            r = total_row + i
            ws_stats.cell(row=r, column=1, value=brand)
            ws_stats.cell(row=r, column=date_col, value=group_data[brand]["수량"])
            ws_stats.cell(row=r, column=date_col + 1, value=group_data[brand]["금액"])
        total_row += len(missing)

    # 합 계 업데이트
    if total_row:
        total_qty = sum(d["수량"] for d in group_data.values())
        total_amt = sum(d["금액"] for d in group_data.values())
        ws_stats.cell(row=total_row, column=date_col, value=total_qty).font = Font(bold=True)
        ws_stats.cell(row=total_row, column=date_col + 1, value=total_amt).font = Font(bold=True)


def _write_stats_sheet(ws_stats, ref_date: date, brand_group_totals: dict,
                       group_cols: list, all_brands: list):
    """
    통계 시트를 부서별 섹션으로 처음부터 새로 작성한다.
    API에서 전체 데이터를 가져오므로 기존 내용은 지우고 재작성한다.

    brand_group_totals: {group: {brand: {수량: int, 금액: int}}}
    all_brands: 브랜드 표시 순서 목록
    """
    # 기존 내용 전체 삭제 후 재작성 (값 + 서식 모두 초기화)
    for row in ws_stats.iter_rows():
        for cell in row:
            cell.value = None
            cell.number_format = "General"

    current_row = 1
    for group in group_cols:
        sect_name = GROUP_SECTION_NAMES.get(group, group)
        group_data = brand_group_totals.get(group, {})
        if not group_data:
            continue
        current_row = _write_section(ws_stats, current_row, sect_name,
                                      ref_date, group_data, all_brands)


def build_excel(inventory: dict, base_date: str, base_excel: Path | None,
                out_path: Path, group_cols: list | None = None):
    """
    메인 Excel 생성 함수.
    inventory : fetch_all_groups() 결과
    base_date : 'YYYYMMDD'
    base_excel: 기존 Excel 경로 (없으면 None)
    out_path  : 저장 경로
    group_cols: 포함할 그룹 열 목록 (None이면 전체)
    """
    if group_cols is None:
        group_cols = list(WH_GROUPS.keys())

    ref_date = datetime.strptime(base_date, "%Y%m%d").date()

    # 기존 파일 로드 or 새 워크북
    if base_excel and base_excel.exists():
        wb = openpyxl.load_workbook(base_excel)
        prod_map, name_price = load_base_excel(base_excel)
        print(f"\n기존 Excel 로드: {base_excel} ({len(prod_map)}개 품목코드, {len(name_price)}개 품목명 단가)")
    else:
        wb = openpyxl.Workbook()
        # 기본 시트 제거
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        prod_map = {}
        name_price = {}
        print("\n새 Excel 파일 생성")

    # 브랜드별 품목 그룹핑
    # NAMED_BRANDS에 없는 브랜드는 모두 기타브랜드로 합산
    unmapped_raw_brands: dict[str, int] = {}  # 기타브랜드로 묶인 원본 브랜드명 카운트
    brand_items: dict[str, list] = defaultdict(list)
    for prod_cd, info in sorted(inventory.items()):
        raw_brand = info["brand"]
        brand = normalize_brand_name(raw_brand)
        # normalize 후에도 NAMED_BRANDS에 없으면 품목명에서 브랜드 추출 시도
        if _nfc(brand) not in {_nfc(b) for b in NAMED_BRANDS}:
            extracted = extract_brand(info.get("prod_des", ""))
            extracted_norm = normalize_brand_name(extracted)
            if _nfc(extracted_norm) in {_nfc(b) for b in NAMED_BRANDS}:
                brand = extracted_norm
            else:
                unmapped_raw_brands[raw_brand] = unmapped_raw_brands.get(raw_brand, 0) + 1
                brand = "기타브랜드"
        existing = prod_map.get(prod_cd, {})
        prod_name = existing.get("name") or info["prod_des"]
        # 단가: 기존 Excel 품목코드 → 품목명 → 품목 마스터(in_price) 순으로 사용
        price = (existing.get("price")
                 or name_price.get(prod_name)
                 or (info["in_price"] if info.get("in_price") else None))
        brand_items[brand].append({
            "name":  prod_name,
            "price": price,
            "groups": info["groups"],
        })

    # 기타브랜드로 분류된 브랜드명 출력 (디버그용)
    if unmapped_raw_brands:
        print("\n  [기타브랜드로 묶인 브랜드 목록] ← 이 중 별도 시트가 필요한 게 있으면 알려주세요:")
        for b, cnt in sorted(unmapped_raw_brands.items(), key=lambda x: -x[1]):
            print(f"    {b!r}: {cnt}개 품목")

    # 브랜드 시트 생성/업데이트
    for brand, items in sorted(brand_items.items()):
        sheet_name = BRAND_SHEET_MAP.get(brand, brand)[:31]
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(title=sheet_name)
        _write_brand_sheet(ws, brand, items, group_cols)
        print(f"  시트 [{sheet_name}]: {len(items)}개 품목")

    # 통계 시트 처리: 그룹별 × 브랜드별 수량/금액 집계
    brand_group_totals: dict[str, dict] = {}
    for group in group_cols:
        group_data = {}
        for brand, items in brand_items.items():
            qty = sum(int(it["groups"].get(group, 0)) for it in items)
            amt = sum(int(it["groups"].get(group, 0)) * _to_int_price(it["price"]) for it in items)
            if qty != 0 or amt != 0:
                group_data[brand] = {"수량": qty, "금액": amt}
        brand_group_totals[group] = group_data

    # 통계 브랜드 표시 순서 (STATS_BRAND_ORDER + 실제 존재하는 나머지)
    all_brands_set = set(brand_items.keys())
    ordered = [b for b in STATS_BRAND_ORDER if b in all_brands_set]
    ordered += sorted(all_brands_set - set(ordered))

    if "통계" in wb.sheetnames:
        ws_stats = wb["통계"]
    else:
        ws_stats = wb.create_sheet(title="통계", index=0)

    _write_stats_sheet(ws_stats, ref_date, brand_group_totals, group_cols, ordered)

    # 저장
    out_path.parent.mkdir(exist_ok=True)
    wb.save(out_path)
    print(f"\n저장 완료: {out_path}")


# ─────────────────────────────────────────
# main
# ─────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="이카운트 재고 → Excel 누적 저장")
    parser.add_argument("--date", default=datetime.now().strftime("%Y%m%d"),
                        help="기준일자 YYYYMMDD (기본: 오늘)")
    parser.add_argument("--base", default=None,
                        help="기존 Excel 파일 경로 (생략 시 output/ 폴더에서 최신 파일 자동 탐색)")
    parser.add_argument("--groups", default=None,
                        help="포함할 그룹, 쉼표 구분 (예: 온라인,영업). 기본: 전체")
    args = parser.parse_args()

    group_cols = args.groups.split(",") if args.groups else None

    print("=" * 60)
    print(f" 이카운트 재고 Excel 누적 저장")
    print("=" * 60)
    print(f"  기준일자: {args.date}")
    print(f"  그룹:     {group_cols or '전체'}")

    # 기존 파일 탐색
    base_excel = None
    if args.base:
        base_excel = Path(args.base)
    else:
        out_dir = Path("output")
        if out_dir.exists():
            # ~$ 로 시작하는 엑셀 잠금 파일 제외
            xlsx_files = sorted(
                [f for f in out_dir.glob("*재고현황*.xlsx") if not f.name.startswith("~$")],
                reverse=True
            )
            if xlsx_files:
                base_excel = xlsx_files[0]
                print(f"  기존 파일: {base_excel} (자동 탐색)")

    print(f"\n[1/2] ECOUNT 재고 조회 시작...")
    try:
        inventory = fetch_all_groups(args.date)
    except Exception as e:
        print(f"\n[실패] 재고 조회 오류: {e}")
        traceback.print_exc()
        return 1

    total_items = len(inventory)
    print(f"\n조회 완료: 총 {total_items}개 품목")

    stamp = datetime.now().strftime("%y%m%d")
    out_path = Path("output") / f"{stamp}_재고현황.xlsx"

    print(f"\n[2/2] Excel 생성 중...")
    try:
        build_excel(inventory, args.date, base_excel, out_path, group_cols)
    except Exception as e:
        print(f"\n[실패] Excel 생성 오류: {e}")
        traceback.print_exc()
        return 1

    print("\n" + "=" * 60)
    print(" ✅ 완료!")
    print("=" * 60)
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception:
        print("\n[예상치 못한 오류]")
        traceback.print_exc()
        sys.exit(1)
