"""
프레젠스월드 재고현황 자동화 스크립트

사용법:
    python inventory_update.py \
        --inventory ESZ018R_7.xlsx \
        --master ESA009M_8.csv \
        --template 프레젠스월드_재고현황.xlsx \
        --date 2026-06-23

    --date: 엑셀에 기록할 날짜 (작업 당일, 보통 월요일)
            자료 기준일(지난주 금요일)과 다를 수 있음
    --output: 출력 파일 경로 (기본: 원본 덮어쓰기)

필요 파일:
    1. ESZ018R 창고별재고현황 xlsx: 이카운트 → 재고1 → 출력물 → 창고별재고현황
    2. ESA009M 품목마스터 csv: 입고단가 보완용 (이카운트에 단가 없을 때)
    3. 프레젠스월드_재고현황.xlsx: 기존 작업 파일
"""

import argparse
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────
# 창고 컬럼 → 부서 매핑
# 이카운트 창고별재고현황 열이름 → 엑셀 "XX 수량" 컬럼 접두사
# ──────────────────────────────────────────────────────────────
WAREHOUSE_TO_DEPT = {
    # 온라인사업부
    "디아이 미판매(온라인사업부)":         "온라인",
    "디아이 판매(온라인사업부)":           "온라인",
    "본사 교환새상품(온라인사업부)":        "온라인",
    "본사 미판매(온라인사업부)":           "온라인",
    "본사 샘플(온라인사업부)":            "온라인",
    "본사 판매(온라인사업부)":            "온라인",
    "조은 판매(온라인사업부)":            "온라인",
    # 영업사업부
    "영업관리팀 디아이 B2B (유통)":       "영업",
    "영업관리팀 본사 (유통)":            "영업",
    "영업팀 디아이(유통)":              "영업",
    "영업팀 본사 (유통)":              "영업",
    "영업팀 조은(유통)":               "영업",
    # 영업_위탁매장
    "신세계백화점강남점(위탁)":           "영업_위탁매장",
    "아트박스(유통)":                  "영업_위탁매장",
    "애니메이트 부산 DP용 (유통)":        "영업_위탁매장",
    "애니메이트 부산 위탁(유통)":         "영업_위탁매장",
    "애니메이트 잠실 DP용 (유통)":        "영업_위탁매장",
    "애니메이트 잠실 위탁(유통)":         "영업_위탁매장",
    "애니메이트 홍대 DP용 (유통)":        "영업_위탁매장",
    "애니메이트 홍대 위탁(유통)":         "영업_위탁매장",
    "애니모어 위탁(유통)":              "영업_위탁매장",
    "애니플러스 위탁(유통)":             "영업_위탁매장",
    "판매 및 교환 가능 샘플 본사(유통)":    "영업_위탁매장",
    "회신대기(유통)":                  "영업_위탁매장",
    # 사업지원TF
    "사업지원TF(디아이)":              "사업지원",
    # 컨텐츠기획사업부
    "디아이 판매(컨텐츠기획사업부)":       "컨기",
    "본사 불량품(컨텐츠기획사업부)":       "컨기",
    "본사 샘플(컨텐츠기획사업부)":        "컨기",
    "본사 판매(컨텐츠기획사업부)":        "컨기",
    # 직영사업부 (메가하우스_직영 시트 - 수동 붙여넣기)
    "메가캠퍼스 매장(디아이)":           "직영",
    "메가캠퍼스 매장(본사)":            "직영",
    "메가캠퍼스 매장(홍대)":            "직영",
    # 아래 창고는 집계에서 제외 (샘플/불량/AGF)
    # "AGF 창고", "DP용 샘플 본사(유통)", "메가캠퍼스 샘플",
    # "반품,불량 디아이 (유통)", "반품,불량 본사 (유통)"
}

# ──────────────────────────────────────────────────────────────
# 브랜드명 → 엑셀 시트명 매핑
# 이카운트 브랜드 → 프레젠스월드_재고현황.xlsx 시트명
# ──────────────────────────────────────────────────────────────
BRAND_TO_SHEET = {
    # 굿스마일 그룹
    "굿스마일":                "굿스마일",
    "굿스마일(good smile)":     "굿스마일",
    "good smile":              "굿스마일",
    "good smile company":      "굿스마일",
    "goodsmilecompany":        "굿스마일",
    "good smile arts shanghai": "굿스마일",
    "goodsmile racing":        "굿스마일",
    "goodsmile moment":        "굿스마일",
    "good smile racing":       "굿스마일",
    "goodsmile racing":        "굿스마일",
    "max factory":             "굿스마일",
    "orange rouge":            "굿스마일",
    "freeing":                 "굿스마일",
    "freeing":                 "굿스마일",
    "phat!":                   "굿스마일",
    "phat! company":           "굿스마일",
    # 메가하우스
    "메가하우스":               "메가하우스",
    "메가하우스(megahouse)":     "메가하우스",
    "megahouse":               "메가하우스",
    # 부시로드
    "부시로드":                 "부시로드",
    "bushiroad creative":      "부시로드",
    # 블로키
    "블로키":                  "블로키 ",
    # 반다이남코
    "반다이남코":               "반다이남코",
    "반다이":                  "반다이남코",
    "bandai namco arts":       "반다이남코",
    "bandai namco filmworks":  "반다이남코",
    # 핫토이
    "핫토이":                  "핫토이_온라인",
    "핫토이(hottoys)":          "핫토이_온라인",
    "hot toys":               "핫토이_온라인",
    # 하비재팬
    "하비재팬":                "하비재팬",
    # CCS TOYS
    "ccs toys":               "CCS ",
    # 코코파
    "코코파":                  "코코파",
    "코코파스튜디오":            "코코파",
    # 미쿠감사제
    "미쿠감사제_2026":          "미쿠감사제",
    # P001 (품목명으로 판단 - 브랜드가 다를 수 있음)
    "p001":                   "P001 온라인",
}

# 각 시트의 부서 수량 컬럼명 (순서 중요)
SHEET_QTY_COLS = {
    "굿스마일":      ["온라인 수량", "영업 수량", "영업_위탁매장 수량", "사업지원 수량"],
    "메가하우스":    ["온라인 수량", "영업 수량", "사업지원 수량"],
    "부시로드":     ["온라인 수량", "영업 수량", "영업_위탁매장 수량", "사업지원 수량"],
    "블로키 ":      ["온라인 수량", "영업 수량"],
    "반다이남코":   ["온라인 수량", "영업 수량"],
    "핫토이_온라인": ["온라인 수량"],
    "하비재팬":     ["온라인 수량"],
    "CCS ":        ["온라인 수량"],
    "P001 온라인":  ["온라인 수량"],
    "코코파":       ["온라인 수량"],
    "기타브랜드":   ["온라인 수량", "영업 수량"],
    "미쿠감사제":   ["컨기 수량"],
}

# 통계 시트 참조 수식 마지막 행 번호 (시트별)
SHEET_LAST_ROW = {
    "굿스마일":      None,  # 런타임에 감지
    "메가하우스":    None,
    "부시로드":     None,
    "블로키 ":      None,
    "반다이남코":   None,
    "핫토이_온라인": None,
    "하비재팬":     None,
    "CCS ":        None,
    "P001 온라인":  None,
    "코코파":       None,
    "기타브랜드":   None,
    "미쿠감사제":   None,
}


# ──────────────────────────────────────────────────────────────
# 헬퍼 함수
# ──────────────────────────────────────────────────────────────

def normalize_brand(raw: str) -> str:
    key = raw.lower().strip()
    return BRAND_TO_SHEET.get(key, "기타브랜드")


def int_or_zero(v) -> int:
    if v is None:
        return 0
    try:
        return int(v)
    except (ValueError, TypeError):
        return 0


def parse_args():
    p = argparse.ArgumentParser(description="프레젠스월드 재고현황 자동 업데이트")
    p.add_argument("--inventory", required=True, help="이카운트 창고별재고현황 xlsx 경로 (ESZ018R)")
    p.add_argument("--master",    required=True, help="이카운트 ESA009M 품목마스터 csv 경로")
    p.add_argument("--template",  required=True, help="프레젠스월드_재고현황.xlsx 경로")
    p.add_argument("--date", help="엑셀 기록 날짜 YYYY-MM-DD (기본: 오늘)")
    p.add_argument("--output", help="출력 파일 경로 (기본: 원본 덮어쓰기)")
    return p.parse_args()


# ──────────────────────────────────────────────────────────────
# 1) ESA009M 품목마스터 로드 → {품목명: 입고단가}
# ──────────────────────────────────────────────────────────────

def load_price_master(csv_path: str) -> dict:
    prices = {}
    with open(csv_path, encoding="utf-8-sig") as f:
        lines = f.readlines()
    for line in lines[2:]:
        line = line.strip()
        if not line:
            continue
        parts = re.split(r'"\s*,\s*"', line.strip('"'))
        if len(parts) < 4:
            continue
        name = parts[2].strip().rstrip("\t")
        price_str = parts[3].strip().rstrip("\t").replace(",", "")
        try:
            prices[name] = int(price_str)
        except ValueError:
            pass
    return prices


# ──────────────────────────────────────────────────────────────
# 2) ESZ018R 창고별재고현황 로드 → 시트별 품목 데이터
# ──────────────────────────────────────────────────────────────

def load_inventory(xlsx_path: str) -> tuple[dict, set]:
    """
    반환: (by_sheet, unmapped_brands)
      by_sheet: {시트명: [{품목명, 입고단가, 온라인:N, 영업:N, ...}]}
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    # 헤더 행 (2행)
    headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    col_idx = {h: i for i, h in enumerate(headers) if h}  # 0-based index

    # 창고 컬럼 인덱스 목록
    wh_cols = {wh: col_idx[wh] for wh in WAREHOUSE_TO_DEPT if wh in col_idx}

    by_sheet: dict[str, dict] = {}  # {sheet: {품목명: {dept: qty, 입고단가: N}}}
    unmapped_brands = set()
    current_brand = None

    for r in range(3, ws.max_row + 1):
        brand_val = ws.cell(r, col_idx.get("브랜드", 1) + 1).value
        if brand_val:
            current_brand = brand_val.strip()

        item_name = ws.cell(r, col_idx.get("품목명", 3) + 1).value
        if not item_name:
            continue
        item_name = item_name.strip()

        price_raw = ws.cell(r, col_idx.get("입고단가", 4) + 1).value
        try:
            price = int(price_raw) if price_raw else None
        except (ValueError, TypeError):
            price = None

        sheet_name = normalize_brand(current_brand or "")
        if sheet_name == "기타브랜드" and current_brand:
            unmapped_brands.add(current_brand)

        if sheet_name not in by_sheet:
            by_sheet[sheet_name] = {}

        # 이미 같은 품목명이 있으면 수량 합산 (중복 행 처리)
        if item_name not in by_sheet[sheet_name]:
            by_sheet[sheet_name][item_name] = {
                "입고단가": price,
                "브랜드": current_brand or "",
                "온라인": 0, "영업": 0, "영업_위탁매장": 0,
                "사업지원": 0, "컨기": 0, "직영": 0,
            }
        elif price and not by_sheet[sheet_name][item_name]["입고단가"]:
            by_sheet[sheet_name][item_name]["입고단가"] = price

        item = by_sheet[sheet_name][item_name]
        for wh, col_0 in wh_cols.items():
            v = int_or_zero(ws.cell(r, col_0 + 1).value)
            dept = WAREHOUSE_TO_DEPT[wh]
            item[dept] = item.get(dept, 0) + v

    return by_sheet, unmapped_brands


# ──────────────────────────────────────────────────────────────
# 3) 브랜드 시트 업데이트
# ──────────────────────────────────────────────────────────────

def update_brand_sheet(ws, sheet_name: str, items: dict, master_prices: dict):
    """
    ws: 브랜드 시트 (openpyxl worksheet)
    items: {품목명: {dept: qty, 입고단가: N, 브랜드: str}}
    """
    # 헤더 컬럼 위치
    header = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v:
            header[v] = c

    qty_cols = SHEET_QTY_COLS.get(sheet_name, ["온라인 수량"])

    # 기존 품목명 → 행 번호
    name_col = header.get("품목명", 2)
    existing: dict[str, int] = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, name_col).value
        if v:
            existing[str(v).strip()] = r

    price_col = header.get("입고단가")
    brand_col = header.get("브랜드", 1)

    # 기존 수량 컬럼 0으로 초기화
    for col_name in qty_cols:
        cidx = header.get(col_name)
        if cidx:
            for r in range(2, ws.max_row + 1):
                ws.cell(r, cidx).value = 0

    updated_count = new_count = 0

    for item_name, data in items.items():
        # 입고단가: 이카운트 우선, 없으면 품목마스터 참조
        price = data["입고단가"] or master_prices.get(item_name)

        # 각 부서 수량
        qty_map = {}
        for col_name in qty_cols:
            dept_prefix = col_name.replace(" 수량", "")
            qty_map[col_name] = data.get(dept_prefix, 0)

        if item_name in existing:
            r = existing[item_name]
            if price and price_col:
                ws.cell(r, price_col).value = price
            for col_name, qty in qty_map.items():
                cidx = header.get(col_name)
                if cidx:
                    ws.cell(r, cidx).value = qty
            updated_count += 1
        else:
            # 이 시트에서 추적하는 부서 수량이 모두 0이면 신규 추가 안 함
            if all(v == 0 for v in qty_map.values()):
                continue
            # 신규 행 추가
            r = ws.max_row + 1
            # 브랜드: 이 시트의 첫 데이터 행에서 가져옴
            ws.cell(r, brand_col).value = ws.cell(2, brand_col).value
            ws.cell(r, name_col).value = item_name
            if price and price_col:
                ws.cell(r, price_col).value = price
            for col_name, qty in qty_map.items():
                cidx = header.get(col_name)
                if cidx:
                    ws.cell(r, cidx).value = qty
            # 재고금액 수식
            for col_name in list(header.keys()):
                if col_name and "재고금액" in str(col_name):
                    amt_cidx = header[col_name]
                    qty_col_name = col_name.replace("재고금액", "수량")
                    qty_cidx = header.get(qty_col_name)
                    if qty_cidx and price_col:
                        ws.cell(r, amt_cidx).value = (
                            f"={get_column_letter(qty_cidx)}{r}"
                            f"*{get_column_letter(price_col)}{r}"
                        )
            new_count += 1

    return updated_count, new_count


# ──────────────────────────────────────────────────────────────
# 4) 통계 시트 — 새 날짜 컬럼 추가
# ──────────────────────────────────────────────────────────────

def add_stats_column(wb: openpyxl.Workbook, new_date: datetime):
    ws = wb["통계"]

    # 마지막 날짜 컬럼 위치 찾기
    last_date_col = 1
    for c in range(1, ws.max_column + 1):
        if isinstance(ws.cell(1, c).value, datetime):
            last_date_col = c

    new_col = last_date_col + 2  # 날짜, 수량, 금액 쌍

    ws.cell(1, new_col).value = new_date
    ws.cell(2, new_col).value = "수량"
    ws.cell(2, new_col + 1).value = "금액"

    print(f"\n  [통계] {new_date.strftime('%Y-%m-%d')} → {get_column_letter(new_col)}열 추가")

    # 기존 수식 패턴으로부터 참조 시트/셀 복제
    prev_qty_col = last_date_col
    prev_amt_col = last_date_col + 1

    for r in range(3, ws.max_row + 1):
        prev_qty = ws.cell(r, prev_qty_col).value
        prev_amt = ws.cell(r, prev_amt_col).value

        if isinstance(prev_qty, str) and prev_qty.startswith("="):
            # 수식에서 시트!셀 참조 추출 후 마지막 행 번호로 교체
            new_qty_formula = _update_formula_last_row(prev_qty, wb)
            ws.cell(r, new_col).value = new_qty_formula
        elif isinstance(prev_qty, str) and prev_qty.startswith("=SUM"):
            # =SUM(H3:H13) 형태 → 새 컬럼으로 이동
            ws.cell(r, new_col).value = re.sub(
                r"[A-Z]+(\d+:[A-Z]+\d+)",
                lambda m: f"{get_column_letter(new_col)}{m.group(1).split(':')[0][len(get_column_letter(prev_qty_col)):]}"
                          f":{get_column_letter(new_col)}{m.group(1).split(':')[1][len(get_column_letter(prev_qty_col)):]}",
                prev_qty,
            )

        if isinstance(prev_amt, str) and prev_amt.startswith("="):
            new_amt_formula = _update_formula_last_row(prev_amt, wb)
            ws.cell(r, new_col + 1).value = new_amt_formula
        elif isinstance(prev_amt, str) and prev_amt.startswith("=SUM"):
            ws.cell(r, new_col + 1).value = re.sub(
                r"[A-Z]+",
                get_column_letter(new_col + 1),
                prev_amt,
            )


def _update_formula_last_row(formula: str, wb: openpyxl.Workbook) -> str:
    """='시트명'!D1242 형태에서 실제 마지막 데이터 행으로 교체"""
    match = re.match(r"='?([^'!]+)'?!([A-Z]+)(\d+)", formula)
    if not match:
        return formula
    sheet_name, col_letter, old_row = match.groups()
    if sheet_name in wb.sheetnames:
        ws_ref = wb[sheet_name]
        last_row = ws_ref.max_row
        return f"='{sheet_name}'!{col_letter}{last_row}"
    return formula


# ──────────────────────────────────────────────────────────────
# 5) 전주 대비 변동 리포트
# ──────────────────────────────────────────────────────────────

def print_change_report(ws_stats, new_date: datetime):
    print("\n" + "=" * 65)
    print(f"  전주 대비 재고 변동 리포트  ({new_date.strftime('%Y-%m-%d')} 기준)")
    print("=" * 65)

    # 날짜가 있는 컬럼 수집
    date_cols = []
    for c in range(1, ws_stats.max_column + 1):
        if isinstance(ws_stats.cell(1, c).value, datetime):
            date_cols.append(c)

    if len(date_cols) < 2:
        print("  (비교할 이전 주 데이터 없음)")
        return

    prev_date_col = date_cols[-2]
    curr_date_col = date_cols[-1]
    prev_date = ws_stats.cell(1, prev_date_col).value
    curr_date = ws_stats.cell(1, curr_date_col).value

    # 각 섹션 합계 행 비교
    section_label = ""
    for r in range(1, ws_stats.max_row + 1):
        v = ws_stats.cell(r, 1).value
        if isinstance(v, str) and v in ("온라인사업부", "영업사업부", "영업사업부_위탁매장", "사업지원 TF", "컨텐츠기획부", "직영사업부"):
            section_label = v
            continue
        if not isinstance(v, str):
            continue
        if "합 계" not in v:
            continue

        prev_qty = ws_stats.cell(r, prev_date_col).value
        curr_qty = ws_stats.cell(r, curr_date_col).value
        prev_amt = ws_stats.cell(r, prev_date_col + 1).value
        curr_amt = ws_stats.cell(r, curr_date_col + 1).value

        pq = prev_qty if isinstance(prev_qty, (int, float)) else 0
        cq = curr_qty if isinstance(curr_qty, (int, float)) else 0
        pa = prev_amt if isinstance(prev_amt, (int, float)) else 0
        ca = curr_amt if isinstance(curr_amt, (int, float)) else 0

        diff_q = cq - pq
        diff_a = ca - pa
        pct_q = (diff_q / pq * 100) if pq else 0

        marker = " ◀ 주목" if abs(pct_q) >= 10 else ""
        print(f"\n  [{section_label}]")
        print(f"    수량: {pq:>8,.0f} → {cq:>8,.0f}  ({diff_q:+,.0f}, {pct_q:+.1f}%){marker}")
        print(f"    금액: {pa:>14,.0f} → {ca:>14,.0f}  ({diff_a:+,.0f})")

    print("\n" + "=" * 65)


# ──────────────────────────────────────────────────────────────
# main
# ──────────────────────────────────────────────────────────────

def main():
    args = parse_args()
    work_date = (
        datetime.strptime(args.date, "%Y-%m-%d") if args.date else datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)
    )

    print(f"[1/4] 품목마스터 로딩: {args.master}")
    master_prices = load_price_master(args.master)
    print(f"      → {len(master_prices):,}개 품목")

    print(f"\n[2/4] 창고별재고현황 로딩: {args.inventory}")
    by_sheet, unmapped_brands = load_inventory(args.inventory)
    total_items = sum(len(v) for v in by_sheet.values())
    print(f"      → 총 {total_items:,}개 품목 / {len(by_sheet)}개 시트")
    if unmapped_brands:
        print(f"      ※ 기타브랜드로 분류: {sorted(unmapped_brands)}")

    print(f"\n[3/4] 엑셀 업데이트: {args.template}")
    wb = openpyxl.load_workbook(args.template)

    total_updated = total_new = 0
    for sheet_name, items in by_sheet.items():
        if sheet_name not in wb.sheetnames:
            print(f"      [SKIP] 시트 없음: '{sheet_name}'")
            continue
        ws = wb[sheet_name]
        updated, new = update_brand_sheet(ws, sheet_name, items, master_prices)
        total_updated += updated
        total_new += new
        print(f"      [{sheet_name:12s}] 업데이트 {updated:4d}건 / 신규 {new:3d}건")

    add_stats_column(wb, work_date)

    output_path = args.output or args.template
    wb.save(output_path)

    print(f"\n[4/4] 저장: {output_path}")
    print(f"      업데이트 {total_updated}건 / 신규 추가 {total_new}건")

    print_change_report(wb["통계"], work_date)
    print(f"\n완료. 직영사업부(메가하우스_직영 시트)는 메일 수신 자료를 수동으로 붙여넣기 해주세요.")


if __name__ == "__main__":
    main()
