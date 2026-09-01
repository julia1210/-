"""
ESZ018R 창고별재고현황 파일에서 브랜드명 목록을 추출합니다.
사용법: python debug_brands.py 창고별재고현황.xlsx
"""
import sys
import openpyxl

path = sys.argv[1] if len(sys.argv) > 1 else "창고별재고현황.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb.active

headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
col_idx = {h: i for i, h in enumerate(headers) if h}

brand_col = col_idx.get("브랜드", 1) + 1  # 1-based

brands = []
for r in range(3, ws.max_row + 1):
    v = ws.cell(r, brand_col).value
    if v and str(v).strip():
        brands.append(str(v).strip())

print("ESZ018R 파일의 브랜드명 목록:")
for b in sorted(set(brands)):
    print(f"  {repr(b)}")
